import os
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash
from openpyxl import load_workbook

load_dotenv()

from app import create_app, db
from app.models import User, Slot

DEFAULT_PASSWORD = os.getenv("DEFAULT_PASSWORD", "ChangeMe123!")

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "ELENCO UNICO CORSO.xlsx")

def slug_email(nome: str, cognome: str) -> str:
    base = f"{nome}.{cognome}".strip().lower()
    base = base.replace(" ", "").replace("'", "").replace("’", "")
    base = base.replace("à","a").replace("è","e").replace("é","e").replace("ì","i").replace("ò","o").replace("ù","u")
    return f"{base}@smam.local"

def detect_columns(headers):
    # ritorna indici colonna per nome/cognome/gruppo
    h = [str(x).strip().lower() if x is not None else "" for x in headers]

    def find_any(cands):
        for idx, val in enumerate(h):
            for c in cands:
                if c in val:
                    return idx
        return None

    col_nome = find_any(["nome"])
    col_cognome = find_any(["cognome"])
    # gruppo può chiamarsi anche plotone/classe ecc.
    col_gruppo = find_any(["gruppo", "plotone", "classe"])
    return col_nome, col_cognome, col_gruppo

def load_users_from_excel(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel not found: {path}")

    wb = load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_nome, col_cognome, col_gruppo = detect_columns(headers)

    if col_nome is None or col_cognome is None:
        raise ValueError("Nel file Excel non trovo le colonne 'Nome' e/o 'Cognome' nella prima riga.")

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[col_nome]).strip() if row[col_nome] else ""
        cognome = str(row[col_cognome]).strip() if row[col_cognome] else ""
        gruppo = str(row[col_gruppo]).strip() if (col_gruppo is not None and row[col_gruppo]) else ""

        if not nome or not cognome:
            continue

        users.append((nome, cognome, gruppo))

    return users

def seed_slots():
    # PALESTRA Lun–Ven: 1)16:00-17:15 2)17:15-18:15 3)20:00-21:15 cap 30
    for dow in [1,2,3,4,5]:
        db.session.add(Slot(impianto="PALESTRA", titolo="1° Turno", giorno_settimana=dow,
                            ora_inizio="16:00", ora_fine="17:15", capienza=30, attivo=True))
        db.session.add(Slot(impianto="PALESTRA", titolo="2° Turno", giorno_settimana=dow,
                            ora_inizio="17:15", ora_fine="18:15", capienza=30, attivo=True))
        db.session.add(Slot(impianto="PALESTRA", titolo="3° Turno", giorno_settimana=dow,
                            ora_inizio="20:00", ora_fine="21:15", capienza=30, attivo=True))

    # CAMPI Lun–Ven: 16:00–18:15 illimitata
    for dow in [1,2,3,4,5]:
        db.session.add(Slot(impianto="CAMPI", titolo="Unico turno", giorno_settimana=dow,
                            ora_inizio="16:00", ora_fine="18:15", capienza=None, attivo=True))

    # PISCINA:
    # Martedì 17:10–18:00 cap 21 (dow=2)
    db.session.add(Slot(impianto="PISCINA", titolo="Turno", giorno_settimana=2,
                        ora_inizio="17:10", ora_fine="18:00", capienza=21, attivo=True))
    # Mercoledì 16:20–17:10 cap 14 (dow=3)
    db.session.add(Slot(impianto="PISCINA", titolo="Turno 1", giorno_settimana=3,
                        ora_inizio="16:20", ora_fine="17:10", capienza=14, attivo=True))
    # Mercoledì 17:10–18:00 cap 14 (dow=3)
    db.session.add(Slot(impianto="PISCINA", titolo="Turno 2", giorno_settimana=3,
                        ora_inizio="17:10", ora_fine="18:00", capienza=14, attivo=True))
    # Giovedì 17:10–18:00 cap 21 (dow=4)
    db.session.add(Slot(impianto="PISCINA", titolo="Turno", giorno_settimana=4,
                        ora_inizio="17:10", ora_fine="18:00", capienza=21, attivo=True))

def ensure_single_admin():
    # forza: solo Meneghelli Massimo (ATLA)
    admin_nome = "Massimo"
    admin_cognome = "Meneghelli"
    admin_gruppo = "ATLA"
    admin_email = slug_email(admin_nome, admin_cognome)

    # tutti gli altri -> USER
    User.query.update({User.ruolo: "USER"})
    db.session.flush()

    admin = User.query.filter_by(email=admin_email).first()
    if not admin:
        admin = User(
            nome=admin_nome,
            cognome=admin_cognome,
            gruppo=admin_gruppo,
            ruolo="ADMIN",
            email=admin_email,
            password_hash=generate_password_hash(DEFAULT_PASSWORD),
        )
        db.session.add(admin)
    else:
        admin.nome = admin_nome
        admin.cognome = admin_cognome
        admin.gruppo = admin_gruppo
        admin.ruolo = "ADMIN"
        # non tocchiamo password se già esiste
    db.session.flush()

def main():
    app = create_app()
    with app.app_context():
        db.create_all()

        # Seed slots solo se non esistono
        if Slot.query.count() == 0:
            seed_slots()

        # Import utenti Excel (aggiunge solo quelli mancanti)
        imported = 0
        try:
            rows = load_users_from_excel(EXCEL_PATH)
        except Exception as e:
            print(f"[WARN] Impossibile importare utenti da Excel: {e}")
            rows = []

        for nome, cognome, gruppo in rows:
            email = slug_email(nome, cognome)
            exists = User.query.filter_by(email=email).first()
            if exists:
                # aggiorno campi base senza toccare password
                exists.nome = nome
                exists.cognome = cognome
                exists.gruppo = gruppo
                continue

            u = User(
                nome=nome,
                cognome=cognome,
                gruppo=gruppo,
                ruolo="USER",
                email=email,
                password_hash=generate_password_hash(DEFAULT_PASSWORD),
            )
            db.session.add(u)
            imported += 1

        ensure_single_admin()

        db.session.commit()
        print(f"[OK] DB ready. Imported new users: {imported}. Slots: {Slot.query.count()}, Users: {User.query.count()}")

if __name__ == "__main__":
    main()
